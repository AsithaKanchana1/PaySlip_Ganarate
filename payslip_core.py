"""
payslip_core.py
===============
Core pay slip generation logic — no GUI dependencies.
Imported by app.py (GUI) and generate_payslips.py (CLI).
"""

import openpyxl
import io
import re
try:
    import msoffcrypto
except Exception:
    msoffcrypto = None
try:
    import xlrd
except Exception:
    xlrd = None
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
import os

# ─── Column indices (0-based, matching Excel sheet) ───────────────────────────
COL_EMP_NO        = 0
COL_NAME          = 1
COL_DEPARTMENT    = 2
COL_DESIGNATION   = 3
COL_BASIC         = 4
COL_ALLOWANCE     = 5
COL_ATT_BONUS     = 6
COL_FIXED_ALLOW   = 7
COL_INCENTIVE     = 8
COL_NORMAL_OT     = 9
COL_NORMAL_OT_AMT = 10
COL_DOUBLE_OT     = 11
COL_DOUBLE_OT_AMT = 12
COL_INCENTIVE2    = 13
COL_GROSS         = 14
COL_SAL_ADV       = 15
COL_NO_PAY        = 16
COL_NO_PAY_AMT    = 17
COL_ATT_BON_DED   = 18
COL_ALLOW_DED     = 19
COL_EPF8          = 20
COL_LATE          = 21
COL_LATE_DED      = 22
COL_WELFARE       = 23
COL_TOTAL_DED     = 24
COL_NET_SALARY    = 25
COL_EPF12         = 26
COL_ETF3          = 27

# ─── Page layout constants ────────────────────────────────────────────────────
# 4 slips per page (2 columns × 2 rows), each slip sized to fill A4.
# PAGE_MARGIN: small safe-area margin for printers (all four sides).
# GAP:         gap between adjacent slips (horizontal and vertical).
COLS_PER_PAGE = 2
ROWS_PER_PAGE = 2
PAGE_MARGIN   = 5  * mm   # 5 mm on all sides
GAP           = 2  * mm   # 2 mm between slips

# Slip dimensions — computed from A4 minus margins and gaps so the slips
# fill the page exactly. Result ≈ 99 mm × 142.5 mm per slip.
_A4_W, _A4_H  = A4
SLIP_WIDTH  = (_A4_W - 2 * PAGE_MARGIN - (COLS_PER_PAGE - 1) * GAP) / COLS_PER_PAGE
SLIP_HEIGHT = (_A4_H - 2 * PAGE_MARGIN - (ROWS_PER_PAGE - 1) * GAP) / ROWS_PER_PAGE

# ─── Colours ──────────────────────────────────────────────────────────────────
C_HDR_BG      = colors.HexColor("#1a2744")
C_HDR_TEXT    = colors.white
C_BORDER      = colors.HexColor("#2d3a6b")
C_BORDER_LITE = colors.HexColor("#c8d0e8")
C_ROW_ALT     = colors.HexColor("#f4f7fc")
C_GROSS_BG    = colors.HexColor("#fff8e1")
C_DED_BG      = colors.HexColor("#fde8e8")
C_NET_BG      = colors.HexColor("#e8f8ee")
C_EPF_BG      = colors.HexColor("#f0f0f0")
C_TEXT        = colors.HexColor("#12192b")
C_MUTED       = colors.HexColor("#5a6080")
C_ACCENT      = colors.HexColor("#2d3a8c")


# ─── Helpers ──────────────────────────────────────────────────────────────────

def read_employees(excel_path: str) -> list:
    """Return list of rows (each a list of cell values), skipping header row.

    This function is robust to both .xls and .xlsx formats and can handle
    password-protected files if they are decrypted first using msoffcrypto.
    If the workbook contains a header row somewhere near the top, we try
    to detect it automatically and map necessary columns from the template
    into the internal column layout used by the generator.
    """
    # For backward compatibility, keep a simple call signature where the
    # caller may optionally pass a tuple (path, password) encoded as
    # "path::password". This keeps API simple for callers that want to
    # provide a password via a single string. Otherwise, they should pass
    # the plain path and manage decryption externally.
    password = None
    if isinstance(excel_path, str) and "::" in excel_path:
        excel_path, password = excel_path.split("::", 1)

    def _read_all_rows_from_stream(stream, ext):
        rows = []
        if ext == ".xls":
            if xlrd is None:
                raise RuntimeError("xlrd is required to read .xls files (install xlrd>=2.0.1)")
            data = stream.read()
            book = xlrd.open_workbook(file_contents=data)
            sheet = book.sheet_by_index(0)
            for r in range(sheet.nrows):
                row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                rows.append(row)
        else:
            # .xlsx or unknown - let openpyxl handle it
            stream.seek(0)
            wb = openpyxl.load_workbook(stream, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()
        return rows

    ext = os.path.splitext(excel_path)[1].lower()
    # Handle plain files or decrypt if password provided
    if password and msoffcrypto is None:
        raise RuntimeError("msoffcrypto-tool is required to open password-protected files")

    if password:
        with open(excel_path, "rb") as fh:
            of = msoffcrypto.OfficeFile(fh)
            of.load_key(password=password)
            bio = io.BytesIO()
            of.decrypt(bio)
            bio.seek(0)
            rows = _read_all_rows_from_stream(bio, ext)
    else:
        # No password — try to read directly
        if ext == ".xls":
            if xlrd is None:
                raise RuntimeError("xlrd is required to read .xls files")
            book = xlrd.open_workbook(excel_path)
            sheet = book.sheet_by_index(0)
            rows = []
            for r in range(sheet.nrows):
                row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                rows.append(row)
        else:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()

    # Normalize cell values to simple strings/numbers
    def _cell_text(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        try:
            return v
        except Exception:
            return str(v)

    # Detect header row by looking for known header keywords
    header_keywords = [
        "emp", "employee", "name", "basic", "gross", "net",
        "allowance", "attendance", "incentive", "ot", "salary",
    ]

    header_idx = None
    max_scan = min(40, len(rows))
    for i in range(max_scan):
        row = rows[i]
        text = " ".join([str(_cell_text(c)).lower() for c in row if c is not None])
        matches = sum(1 for k in header_keywords if k in text)
        if matches >= 3:
            header_idx = i
            break

    if header_idx is None:
        header_idx = 0

    header_row = [str(_cell_text(c)).strip() for c in rows[header_idx]]

    # Mapping of signature strings -> target internal column constants
    synonyms = {
        # Employee identifier — match whole words only to avoid substring hits (e.g. 'PAID' contains 'id')
        "emp no": COL_EMP_NO, "employee no": COL_EMP_NO, "emp": COL_EMP_NO,
        "name": COL_NAME,
        "department": COL_DEPARTMENT, "dept": COL_DEPARTMENT,
        "designation": COL_DESIGNATION, "desig": COL_DESIGNATION,
        "basic": COL_BASIC,
        "allowance": COL_ALLOWANCE,
        # 'Att. Bonus' normalises to 'att bonus'; also accept plain 'att' or 'attendance'
        "attendance bonus": COL_ATT_BONUS, "att bonus": COL_ATT_BONUS,
        "att": COL_ATT_BONUS, "attendance": COL_ATT_BONUS,
        "fixed allowance": COL_FIXED_ALLOW,
        "incentive": COL_INCENTIVE,
        "normal ot": COL_NORMAL_OT, "normal ot amount": COL_NORMAL_OT_AMT,
        "double ot": COL_DOUBLE_OT, "double ot amount": COL_DOUBLE_OT_AMT,
        "gross salary": COL_GROSS, "gross": COL_GROSS,
        "salary advance": COL_SAL_ADV, "salary adv": COL_SAL_ADV,
        "no pay": COL_NO_PAY, "no pay amount": COL_NO_PAY_AMT,
        "attendance bonus ded": COL_ATT_BON_DED, "attendance bonus deduction": COL_ATT_BON_DED,
        "allowance ded": COL_ALLOW_DED, "allowance deduction": COL_ALLOW_DED,
        # EPF/ETF: 'E.P.F. 8%' normalises to 'e p f 8%' (dots → spaces), so match both forms
        "epf 8%": COL_EPF8, "epf 8": COL_EPF8, "e p f 8%": COL_EPF8, "e p f 8": COL_EPF8,
        "late": COL_LATE, "late deduction": COL_LATE_DED,
        "welfare": COL_WELFARE,
        "total deduction": COL_TOTAL_DED, "total ded": COL_TOTAL_DED,
        "net salary": COL_NET_SALARY, "net": COL_NET_SALARY,
        "epf 12%": COL_EPF12, "e p f 12%": COL_EPF12,
        "etf 3%": COL_ETF3, "e t f 3%": COL_ETF3, "etf": COL_ETF3,
    }

    # Helper normalize header text
    def _norm(h):
        if not h:
            return ""
        t = h.lower()
        t = re.sub(r"[^a-z0-9% ]+", " ", t)
        t = re.sub(r"\s+", " ", t).strip()
        return t

    mapped_cols = {}
    dup_count = {}
    for idx, h in enumerate(header_row):
        nh = _norm(h)
        for key, target in synonyms.items():
            if key in nh:
                # handle duplicate header names (Allowance, Incentive, Attendance Bonus)
                count = dup_count.get(target, 0)
                dup_count[target] = count + 1
                # Map first occurrence to primary target; second to deduction/secondary where applicable
                if target == COL_ALLOWANCE and count >= 1:
                    mapped_cols[COL_ALLOW_DED] = idx
                elif target == COL_ATT_BONUS and count >= 1:
                    mapped_cols[COL_ATT_BON_DED] = idx
                elif target == COL_INCENTIVE and count >= 1:
                    mapped_cols[COL_INCENTIVE2] = idx
                else:
                    mapped_cols[target] = idx
                break

    # If Emp No not found, try to assume first column
    if COL_EMP_NO not in mapped_cols and len(header_row) > 0:
        mapped_cols[COL_EMP_NO] = 0

    # ── Department-header detection ────────────────────────────────────────────
    # The salary sheet embeds department names as section-header rows where
    # ONLY the Name cell is manually entered — the sequential employee-number
    # ("No:") and EPF-number columns are always blank on these rows.
    #
    # Formula-proof design:
    #   Old approach (fragile): count non-None cells == 1
    #     → breaks when formula cells cache as 0 instead of None
    #   New approach (robust): check the TWO manually-entered identity columns
    #     (EPF No and sequential No) — these are NEVER formula-driven and will
    #     always be None on a department-header row, regardless of whether
    #     other columns contain formula-cached values like 0.

    name_col_idx   = mapped_cols.get(COL_NAME,   None)
    emp_no_col_idx = mapped_cols.get(COL_EMP_NO, None)

    # Also find the raw sheet column index for the sequential "No:" column.
    # This is distinct from the EPF column and is a plain integer in real data.
    # Heuristic: scan the header row for "no" / "no:" at a low column index.
    _seq_no_col_idx = None
    for _hdr_idx, _hdr_val in enumerate(header_row):
        _nh = re.sub(r"[^a-z0-9]+", "", str(_hdr_val).lower())
        if _nh in ("no", "no:", "empno", "slno", "sno") and _hdr_idx < 5:
            _seq_no_col_idx = _hdr_idx
            break
    # Fall back to emp_no_col_idx if no separate seq-no column found
    if _seq_no_col_idx is None:
        _seq_no_col_idx = emp_no_col_idx

    def _is_dept_header(raw_row: list) -> str | None:
        """
        Return the department name if this raw row is a section-header row.

        A department-header row has:
          • A non-empty Name cell
          • An empty sequential employee-number column  ← manually entered, never formula
          • An empty EPF-number column                  ← manually entered, never formula

        This check survives formula cells returning 0 or any cached value in other
        columns, because EPF and employee-number are always hand-typed, never computed.
        """
        if name_col_idx is None:
            return None

        # Must have a non-empty name
        name_val = raw_row[name_col_idx] if name_col_idx < len(raw_row) else None
        if not name_val or not str(name_val).strip():
            return None

        # Sequential employee number must be blank (manually entered → None on dept rows)
        seq_no = raw_row[_seq_no_col_idx] if (_seq_no_col_idx is not None and _seq_no_col_idx < len(raw_row)) else None
        if seq_no is not None and str(seq_no).strip() not in ("", "0"):
            return None   # real employee row — has a sequence number

        # EPF number must also be blank
        epf_col = 0  # EPF NO is always the first column in this sheet layout
        epf_val = raw_row[epf_col] if epf_col < len(raw_row) else None
        if epf_val is not None and str(epf_val).strip() not in ("", "0"):
            return None   # real employee row — has an EPF number

        return str(name_val).strip()


    max_cols        = max(COL_ETF3, COL_NET_SALARY) + 1
    current_dept    = ""   # tracks the most-recently-seen department header
    employees       = []

    for row in rows[header_idx + 1:]:
        if not row:
            continue

        # ── Check whether this is a department-header row ──────────────────────
        dept_label = _is_dept_header(row)
        if dept_label is not None:
            current_dept = dept_label
            continue          # don't add this row as an employee

        # ── Skip any row where the Name column is empty ────────────────────────
        # This catches: blank spacer rows, summary/total rows, formula-only rows,
        # and any other non-employee row that has no name to print on the payslip.
        name_val = row[name_col_idx] if (name_col_idx is not None and name_col_idx < len(row)) else None
        if not name_val or not str(name_val).strip():
            continue

        # ── Build the normalised row ───────────────────────────────────────────
        nrow = [None] * max_cols
        for target, col_idx in mapped_cols.items():
            if col_idx < len(row):
                nrow[target] = row[col_idx]

        # Inject department from the last seen section-header row
        # Only overwrite if the sheet has no dedicated department column
        if current_dept and (nrow[COL_DEPARTMENT] is None or str(nrow[COL_DEPARTMENT]).strip() == ""):
            nrow[COL_DEPARTMENT] = current_dept

        employees.append(nrow)

    return employees


def _g(row: list, col: int):
    try:
        return row[col]
    except IndexError:
        return None


def _fmt(val) -> str:
    if val is None:
        return "0.00"
    try:
        return f"{float(val):,.2f}"
    except (ValueError, TypeError):
        return "0.00"


def _fmt_units(val) -> str:
    if val is None or val == 0:
        return ""
    try:
        v = float(val)
        return f"{v:.2f}".rstrip("0").rstrip(".")
    except (ValueError, TypeError):
        return ""


# ─── Draw one pay slip ────────────────────────────────────────────────────────

def draw_payslip(c: canvas.Canvas, x: float, y: float, w: float, h: float,
                 row: list, company_name: str, pay_period: str):
    """Draw one payslip with its top-left at (x, y+h) in PDF coordinates.

    All sizes are derived proportionally from w and h so the slip fills
    whatever rectangle it is given, regardless of page layout.
    """
    W, H = w, h

    # ── Scale factors (vs. original reference slip: 5 cm × 12 cm) ────────────
    ws = W / (5.0 * cm)   # width scale  ≈ 1.98 for new 2-column layout
    hs = H / (12.0 * cm)  # height scale ≈ 1.19 for new 2-row layout

    # ── Font sizes — scaled then clamped to sensible bounds ──────────────────
    F_COMPANY = max(8,  min(14, round(7.5  * ws)))
    F_SUBTITLE= max(7,  min(11, round(6.0  * ws)))
    F_EMP_NO  = max(6,  min(10, round(5.2  * ws)))
    F_NAME    = max(7,  min(11, round(5.5  * ws)))
    F_DEPT    = max(5,  min( 9, round(4.8  * ws)))
    F_NORMAL  = max(6,  min(10, round(5.7  * ws)))
    F_BOLD    = max(7,  min(11, round(6.2  * ws)))
    F_SMALL   = max(5,  min( 9, round(5.4  * ws)))
    F_UNITS   = max(5,  min( 9, round(5.2  * ws)))

    # ── Outer border ─────────────────────────────────────────────────────────
    c.setStrokeColor(C_BORDER)
    c.setLineWidth(0.7)
    c.rect(x, y, W, H, stroke=1, fill=0)

    # ── Header block (13.5 % of slip height) ─────────────────────────────────
    hdr_h   = H * 0.135
    pad_x   = W * 0.02          # small horizontal inset inside header
    pad_top = hdr_h * 0.04     # top padding inside header

    c.setFillColor(C_HDR_BG)
    c.rect(x, y + H - hdr_h, W, hdr_h, stroke=0, fill=1)

    # Distribute 6 header items evenly across hdr_h
    # positions as fractions of hdr_h from the top
    fracs = [0.16, 0.33, 0.49, 0.64, 0.79, 0.95]
    tops  = [y + H - hdr_h * f for f in fracs]

    c.setFillColor(C_HDR_TEXT)

    c.setFont("Helvetica-Bold", F_COMPANY)
    c.drawCentredString(x + W / 2, tops[0], company_name)

    c.setFont("Helvetica-Bold", F_SUBTITLE)
    c.drawCentredString(x + W / 2, tops[1], "Pay Sheet")
    c.drawCentredString(x + W / 2, tops[2], pay_period)

    c.setFont("Helvetica", F_EMP_NO)
    emp_no = _g(row, COL_EMP_NO)
    emp_no_str = str(int(emp_no)) if emp_no else "-"
    c.drawString(x + pad_x, tops[3], f"Emp No: {emp_no_str}")

    c.setFont("Helvetica-Bold", F_NAME)
    name = _g(row, COL_NAME) or ""
    c.drawCentredString(x + W / 2, tops[4], name)

    c.setFont("Helvetica", F_DEPT)
    dept  = _g(row, COL_DEPARTMENT)  or ""
    desig = _g(row, COL_DESIGNATION) or ""
    dept_str = f"Dept: {dept}" + (f"   {desig}" if desig else "")
    c.drawString(x + pad_x, tops[5], dept_str)

    # ── Body data rows ────────────────────────────────────────────────────────
    data_rows = [
        ("Basic",            "",                                    _fmt(_g(row, COL_BASIC)),        "earn"),
        ("Allowance",        "",                                    _fmt(_g(row, COL_ALLOWANCE)),    "earn"),
        ("Attendance Bonus", "",                                    _fmt(_g(row, COL_ATT_BONUS)),    "earn"),
        ("Fixed Allowance",  "",                                    _fmt(_g(row, COL_FIXED_ALLOW)),  "earn"),
        ("Incentive",        "",                                    _fmt(_g(row, COL_INCENTIVE)),    "earn"),
        ("Normal OT",        _fmt_units(_g(row, COL_NORMAL_OT)),   _fmt(_g(row, COL_NORMAL_OT_AMT)),"earn"),
        ("Double OT",        _fmt_units(_g(row, COL_DOUBLE_OT)),   _fmt(_g(row, COL_DOUBLE_OT_AMT)),"earn"),
        ("Incentive",        "",                                    _fmt(_g(row, COL_INCENTIVE2)),   "earn"),
        ("Gross Salary",     "",                                    _fmt(_g(row, COL_GROSS)),        "gross"),
        ("Salary Advance",   "",                                    _fmt(_g(row, COL_SAL_ADV)),      "deduct"),
        ("No Pay",           _fmt_units(_g(row, COL_NO_PAY)),      _fmt(_g(row, COL_NO_PAY_AMT)),   "deduct"),
        ("Att. Bonus Ded.",  "",                                    _fmt(_g(row, COL_ATT_BON_DED)),  "deduct"),
        ("Allowance Ded.",   "",                                    _fmt(_g(row, COL_ALLOW_DED)),    "deduct"),
        ("E.P.F 8%",         "",                                    _fmt(_g(row, COL_EPF8)),         "deduct"),
        ("Late",             _fmt_units(_g(row, COL_LATE)),        _fmt(_g(row, COL_LATE_DED)),     "deduct"),
        ("Welfare",          "",                                    _fmt(_g(row, COL_WELFARE)),      "deduct"),
        ("Total Deduction",  "",                                    _fmt(_g(row, COL_TOTAL_DED)),    "total_d"),
        ("Net Salary",       "",                                    _fmt(_g(row, COL_NET_SALARY)),   "net"),
        ("E.P.F. 12%",       "",                                    _fmt(_g(row, COL_EPF12)),        "epf"),
        ("E.T.F. 3%",        "",                                    _fmt(_g(row, COL_ETF3)),         "epf"),
    ]

    BODY_TOP = y + H - hdr_h
    BODY_H   = BODY_TOP - y
    ROW_H    = BODY_H / len(data_rows)

    # Column widths — proportional to slip width
    COL1_W  = W * 0.54    # label column
    COL2_W  = W * 0.16    # units column
    LABEL_X = x + W * 0.025
    UNITS_X = x + COL1_W + COL2_W / 2
    AMT_X   = x + W - W * 0.025

    c.setStrokeColor(C_BORDER)
    c.setLineWidth(0.5)
    c.line(x, BODY_TOP, x + W, BODY_TOP)

    for i, (label, units, amount, style) in enumerate(data_rows):
        ry = BODY_TOP - (i + 1) * ROW_H

        # Row background
        if style == "gross":
            c.setFillColor(C_GROSS_BG)
            c.rect(x + 0.5, ry + 0.3, W - 1, ROW_H - 0.3, stroke=0, fill=1)
        elif style == "total_d":
            c.setFillColor(C_DED_BG)
            c.rect(x + 0.5, ry + 0.3, W - 1, ROW_H - 0.3, stroke=0, fill=1)
        elif style == "net":
            c.setFillColor(C_NET_BG)
            c.rect(x + 0.5, ry + 0.3, W - 1, ROW_H - 0.3, stroke=0, fill=1)
        elif style == "epf":
            c.setFillColor(C_EPF_BG)
            c.rect(x + 0.5, ry + 0.3, W - 1, ROW_H - 0.3, stroke=0, fill=1)
        elif i % 2 == 0:
            c.setFillColor(C_ROW_ALT)
            c.rect(x + 0.5, ry + 0.3, W - 1, ROW_H - 0.3, stroke=0, fill=1)

        # Row divider
        c.setStrokeColor(C_BORDER_LITE)
        c.setLineWidth(0.20)
        c.line(x, ry, x + W, ry)

        # Text baseline — vertically centred in row
        ty   = ry + ROW_H * 0.25
        bold  = style in ("gross", "total_d", "net")
        small = style == "epf"
        fs_label  = F_SMALL if small else F_BOLD   if bold else F_NORMAL
        fs_amount = F_SMALL if small else F_BOLD   if bold else F_NORMAL

        c.setFont("Helvetica-Bold" if bold else "Helvetica", fs_label)
        c.setFillColor(C_MUTED if small else C_ACCENT if bold else C_TEXT)
        c.drawString(LABEL_X, ty, label)

        if units:
            c.setFont("Helvetica", F_UNITS if small else F_UNITS)
            c.setFillColor(C_MUTED)
            c.drawCentredString(UNITS_X, ty, units)

        c.setFont("Helvetica-Bold" if bold else "Helvetica", fs_amount)
        c.setFillColor(C_MUTED if small else C_ACCENT if bold else C_TEXT)
        c.drawRightString(AMT_X, ty, amount)

    # Vertical column dividers
    body_bottom = BODY_TOP - len(data_rows) * ROW_H
    c.setStrokeColor(C_BORDER_LITE)
    c.setLineWidth(0.22)
    c.line(x + COL1_W, body_bottom, x + COL1_W, BODY_TOP)
    c.line(x + COL1_W + COL2_W, body_bottom, x + COL1_W + COL2_W, BODY_TOP)

    # Bottom border line
    c.setStrokeColor(C_BORDER)
    c.setLineWidth(0.7)
    c.line(x, y, x + W, y)


# ─── Main generation function ─────────────────────────────────────────────────

def generate_pdf(excel_path: str, output_path: str,
                 company_name: str, pay_period: str,
                 progress_callback=None, log_callback=None) -> dict:
    """
    Generate the PDF.

    Args:
        excel_path:        Path to the input Excel file.
        output_path:       Path where the PDF will be saved.
        company_name:      Company name printed on each slip.
        pay_period:        Pay period string, e.g. "May-2026".
        progress_callback: Optional callable(percent: int) for progress updates.
        log_callback:      Optional callable(message: str) for log messages.

    Returns:
        dict with keys: success (bool), message (str), employee_count (int),
                        page_count (int), output_path (str)
    """
    def log(msg):
        if log_callback:
            log_callback(msg)

    def progress(pct):
        if progress_callback:
            progress_callback(pct)

    try:
        log(f"📂  Reading Excel file: {os.path.basename(excel_path)}")
        progress(5)

        employees = read_employees(excel_path)
        if not employees:
            return {"success": False, "message": "No employee data found in Excel file.",
                    "employee_count": 0, "page_count": 0, "output_path": ""}

        log(f"✅  Loaded {len(employees)} employee record(s)")
        progress(15)

        PAGE_W, PAGE_H = A4
        cols          = COLS_PER_PAGE
        rows_per_page = ROWS_PER_PAGE
        per_page      = cols * rows_per_page          # always 4
        total_pages   = (len(employees) - 1) // per_page + 1

        log(f"📐  Layout: {cols} columns × {rows_per_page} rows = {per_page} slips/page")
        log(f"📄  Slip size: {SLIP_WIDTH/mm:.1f} mm × {SLIP_HEIGHT/mm:.1f} mm")
        log(f"📄  Total pages: {total_pages}")
        log(f"🖨️  Generating PDF …")
        progress(20)

        c = canvas.Canvas(output_path, pagesize=A4)
        c.setTitle(f"{company_name} Pay Slips – {pay_period}")
        c.setAuthor(f"{company_name} HR System")

        for idx, row in enumerate(employees):
            if idx % per_page == 0 and idx > 0:
                c.showPage()

            slot    = idx % per_page
            col_idx = slot % cols
            row_idx = slot // cols

            slip_x = PAGE_MARGIN + col_idx * (SLIP_WIDTH  + GAP)
            slip_y = PAGE_H - PAGE_MARGIN - (row_idx + 1) * SLIP_HEIGHT - row_idx * GAP

            draw_payslip(c, slip_x, slip_y, SLIP_WIDTH, SLIP_HEIGHT, row, company_name, pay_period)

            pct = 20 + int(((idx + 1) / len(employees)) * 75)
            progress(pct)

        c.save()
        progress(100)
        log(f"✅  PDF saved successfully!")
        log(f"📁  Location: {output_path}")

        return {
            "success":        True,
            "message":        f"Successfully generated {len(employees)} pay slips across {total_pages} page(s).",
            "employee_count": len(employees),
            "page_count":     total_pages,
            "output_path":    output_path,
        }

    except FileNotFoundError as e:
        return {"success": False, "message": f"File not found: {e}",
                "employee_count": 0, "page_count": 0, "output_path": ""}
    except Exception as e:
        return {"success": False, "message": f"Error: {e}",
                "employee_count": 0, "page_count": 0, "output_path": ""}
